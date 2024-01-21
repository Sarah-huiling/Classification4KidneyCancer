'''
train for WSI

'''

# #from setting import parse_opts
# from datasets.brains18 import BrainS18Dataset
# from model import generate_model
# !/usr/bin/env Python
# coding=utf-8
import json
from datetime import datetime
from functools import reduce

import openpyxl
import pandas as pd
from sklearn.metrics import accuracy_score, recall_score, f1_score
import torch
# import numpy as np
import xlrd
import xlwt
from torch import nn
from torch import optim
import os
import random
import glob
import numpy as np
from torchvision.transforms import transforms

import logger
import torch.nn.functional as F
# from torch.optim import lr_scheduler
from torch.utils.data import DataLoader
# import time
# from utils.logger import log
# from scipy import ndimage
from torch.autograd import Variable
from sklearn import metrics as mt
import os

'''data and model'''
# from data_imgMulCla import MyDataset
# from data_img import MyDataset
# from data_WSI import MyDataset
from data_WSI_tumor1vs2 import MyDataset
from modelLib.vit import VisionTransformer, pretrainedViT
# from MobilNetV2_img import mobilenet_v2
from modelLib.MobileNetV2 import mobilenet_v2
from modelLib.pyramidnet import PyramidNet
from modelLib.ResNet import ResNet50, ResNet101, ResNet152
from modelLib.Vgg import *

os.environ['CUDA_LAUNCH_BLOCKING'] = "0, 1"


def train(alexnet_model, train_loader, epoch, train_dict, logger, criterion, use_gpu):
    alexnet_model.train()
    losss = 0
    for iter, batch in enumerate(train_loader):
        torch.cuda.empty_cache()
        if use_gpu:
            inputs = Variable(batch[0].cuda())
            labels = Variable(batch[1].cuda())
        else:
            inputs, labels = Variable(batch['0']), Variable(batch['1'])

        optimizer.zero_grad()

        outputs = alexnet_model(inputs)
        # print(outputs, labels)
        loss = criterion(outputs, labels)
        # acc(outputs, labels)
        # exit()
        loss.backward()
        optimizer.step()
        losss = losss + loss.item()
        # dice0, dice1, dice2, dice3 = dicev(outputs, labels)
        if (iter + 1) % 10 == 0:
            # print(outputs, labels)
            print(
                'Train Epoch: {} [{}/{} ({:.0f}%)]\tLoss: {:.6f}'.format(
                    epoch, iter, len(train_loader),
                    100. * iter / len(train_loader), losss / (iter + 0.000001)))
    train_dict['loss'].append(losss / (iter + 0.000001))
    logger.scalar_summary('train_loss', losss / (iter + 0.000001), epoch)


def adjust_learning_rate(optimizer, epoch):
    """Sets the learning rate to the initial LR decayed by 10 every 30 epochs"""
    lr = 0.001 * (0.1 ** (epoch // 25))
    for param_group in optimizer.param_groups:
        param_group['lr'] = lr


def ODIR_Metrics(pred, target):
    th = 0.5
    gt = target.flatten()
    pr = pred.flatten()

    gt1 = gt[0::2]
    pr_neg = pr[0::2]  # pr_neg 预测为阴性的概率
    gt2 = gt[1::2]
    pr_pos = pr[1::2]  # pr_pos 预测为阳性的概率

    gt_prePob = []
    for i in range(len(gt2)):
        if gt2[i] == 1:
            gt_prePob.append(pr_pos[i])
        if gt2[i] == 0:
            gt_prePob.append(pr_neg[i])
    preLabel = np.zeros(len(gt2))
    preLabel[pr_pos > th] = 1

    print('=' * 20)
    print('gt2.shape', gt2.shape)
    print('pr2.shape', pr_pos.shape)
    # print('pr_pos.shape', len(pr_pos))
    # fpr, tpr, thresholds = mt.roc_curve(gt2, pr_pos, pos_label=1.0)  # 阳性的概率：以1类作为阳性，则输入预测为1的概率
    # roc_auc2 = mt.auc(fpr, tpr)
    kappa = mt.cohen_kappa_score(gt, pr > th)
    print("1：auc值,", mt.roc_auc_score(gt1, pr_neg), 'acc:', mt.accuracy_score(gt1, pr_neg > th))
    print("2：auc值,", mt.roc_auc_score(gt2, pr_pos), 'acc:', mt.accuracy_score(gt2, pr_pos > th))
    # f1 = mt.f1_score(gt, pr > th, average='micro')
    roc_auc = mt.roc_auc_score(gt2, pr_pos)
    return roc_auc, gt2, gt_prePob, pr_neg, pr_pos


def val_test(alexnet_model, val_loader):
    alexnet_model.eval()
    val_loss = 0
    with torch.no_grad():
        p = []
        g = []
        for iter, batch in enumerate(val_loader):
            torch.cuda.empty_cache()
            if use_gpu:
                inputs = Variable(batch[0].cuda())
                labels = Variable(batch[1].cuda())
            else:
                inputs, labels = Variable(batch['0']), Variable(batch['1'])
            outputs = alexnet_model(inputs)
            loss = criterion(outputs, labels)  # loss为CrossEntropyLoss时，model不需要提前softmax outputs
            outputs = torch.softmax(outputs, dim=1)
            outputs = outputs.data.cpu().numpy()
            labels = labels.cpu().numpy()
            for x, y in zip(outputs, labels):
                p.append(x)
                g.append(y)
            val_loss += loss.item()
        auc, gt, gt_prePob, pr_neg, pr_pos = ODIR_Metrics(np.array(p), np.array(g))
    val_loss /= len(val_loader)
    print('\nAverage loss: {:.6f},auc: {:.6f}\n'.format(val_loss, auc))
    return auc, val_loss, gt, gt_prePob, pr_neg, pr_pos


def loadWsiPath(WsiPath):
    tumor1Path = []
    tumor2Path = []
    if len(glob.glob(WsiPath + '/*FH*Tumor.jpg')):
        tumor1Path = glob.glob(WsiPath + '/*FH*Tumor.jpg')
    if len(glob.glob(WsiPath + '/*PR*Tumor.jpg')):
        tumor2Path = glob.glob(WsiPath + '/*PR*Tumor.jpg')
    return tumor1Path, tumor2Path


class WeightedMultilabel(torch.nn.Module):
    def __init__(self, weights: torch.Tensor):
        self.loss = torch.nn.BCEWithLogitsLoss()
        self.weights = weights.unsqueeze()

    def forward(self, outputs, targets):
        return self.loss(outputs, targets) * self.weights


def check_dir(path):
    if not os.path.exists(path):
        os.mkdir(path)
    return path


if __name__ == "__main__":
    # batch_size = 32
    batch_size = 64
    epochs = 60
    lr = 0.001
    momentum = 0.95
    w_decay = 1e-6
    step_size = 20
    gamma = 0.5
    n_class = 2
    use_gpu = torch.cuda.is_available()
    num_gpu = list(range(torch.cuda.device_count()))

    # path = check_dir('/media/zhl/ResearchData/20221025Huaxi-KidneyCancer-WSI/DLResults/Noverlap54PID/Vit_1vs2_vah230616Corrected')
    Root_SavePath = '/media/zhl/ResearchData/20221025Huaxi-KidneyCancer-WSI/DLResults/Noverlap14PID/20230705/1v2'
    path = check_dir(os.path.join(Root_SavePath, 'vgg16_' + str(datetime.now()).split(':')[0] + ':' +
                                  str(datetime.now()).split(':')[1]))
    trainvalpath = check_dir(path + '/trainvalpath/')
    modelPath = check_dir(path + '/model')
    probPath = check_dir(path + '/prob/')
    aucExcelPath = check_dir(path + '/aucExcel')
    LogPath = check_dir(path + '/Log')


    xl_path = '/media/zhl/ResearchData/20221025Huaxi-KidneyCancer-WSI/DLResults/Noverlap14PID/SelectedSpilit/1v2/test_fold.xls'
    reads = xlrd.open_workbook(xl_path)
    test_path = []
    for row in range(reads.sheet_by_index(0).nrows):
        test_path.append(reads.sheet_by_index(0).cell(row, 0).value)
    # ###### 导入某一折的train  #######
    train_path = []
    reads = xlrd.open_workbook(
        '/media/zhl/ResearchData/20221025Huaxi-KidneyCancer-WSI/DLResults/Noverlap14PID/SelectedSpilit/1v2/train_fold.xls')
    for row in range(reads.sheet_by_index(0).nrows):
        train_path.append(reads.sheet_by_index(0).cell(row, 0).value)
    # # train_path = train_path[:1000]
    # # test_path = test_path[:1000]
    train_transform = transforms.Compose([
        # transforms.RandomResizedCrop((128, 128)),
        transforms.RandomHorizontalFlip(),
        transforms.RandomVerticalFlip(),
        transforms.RandomRotation(90),
        transforms.RandomGrayscale(0.1),
        transforms.ColorJitter(0.3, 0.3, 0.3, 0.3),
        transforms.ToTensor()
    ])

    train_da = MyDataset(train_path, transform=None)
    test = MyDataset(test_path, transform=None)
    # val = MyDataset(val_path, transform=False)
    train_loader = DataLoader(train_da, batch_size=batch_size, shuffle=False, num_workers=3,
                              persistent_workers=False, pin_memory=False)
    test_loader = DataLoader(test, batch_size=batch_size, shuffle=False, num_workers=3,
                             persistent_workers=False, pin_memory=False)
    # val_loader = DataLoader(val, batch_size=batch_size, shuffle=False, num_workers=4)

    print('model load...')
    # 模型以及结果保存
    # model_dir = "/ResearchData/2020-7-1山东肿瘤/Mod_resul/Image/grouping_v1.4"  # models4 130(0.74)
    model_dir = modelPath  # models4 130(0.74)
    if not os.path.exists(model_dir):
        os.makedirs(model_dir)
    # modelName = '_mobilev2.pth'
    # preModel = '/media/zhl/ProgramCode/DL_Classification/preModel/mobilenet_v2-b0353104.pth'
    # model = mobilenet_v2(in_c=3, num_classes=2, pretrained=False, dropoutP=0.85)
    # model = mobilenet_v2(in_c=3, num_classes=2, pretrained=False, input_size=512, model_dir=preModel)
    # modelName = '_resNet50.pth'
    # model = ResNet50()
    # modelName = '_PyramidNet.pth'
    # model = PyramidNet(dataset='imagenet', depth=100, alpha=150, num_classes=2)


    '''# # 2D ViT model # #'''
    # preModel = '/media/zhl/ProgramCode/DL_Classification/pretrainedModel/vit_base_patch16_224_in21k.pth'
    # modelName = '_vit.pth'
    # # model = VisionTransformer(in_c=3, num_classes=2, patch_size=32, img_size=512)
    # model = pretrainedViT(pretrained=True, num_classes=2, patch_size=16, preModel=preModel, in_c=3, img_size=224)
    #                       patch_size=32, img_size=224,drop_ratio=0.5, attn_drop_ratio=0.5, drop_path_ratio=0.2)

    modelName = '_vgg16_bn.pth'
    model = vgg16_bn()
    # model = resnet50(sample_input_D=16, sample_input_H=96, sample_input_W=96, num_seg_classes=2)
    # model = _3DCNN()

    if use_gpu:
        alexnet_model = model.cuda()
        alexnet_model = nn.DataParallel(alexnet_model, device_ids=num_gpu)
    else:
        alexnet_model = model
    # print(model)
    # exit()
    # weight = torch.FloatTensor([0.1, 9]).cuda() # 0.6
    # criterion = nn.BCELoss()  # BCEloss计算前需要将output sigmoid
    # criterion = nn.BCELoss(weight=weight)
    criterion = nn.CrossEntropyLoss()  # label_smoothing: soft label # 计算loss前不需要sigmoid
    # criterion = nn.CrossEntropyLoss(label_smoothing=0.1)  # label_smoothing: soft label
    # criterion = nn.BCEWithLogitsLoss(pos_weight=pos_weight)  # pos_weight：a weight of positive examples.
    #       Must be a vector with length equal to the number of classes，二分类即每张图一个标签，只需要将正样本loss的权重写上即可
    # criterion = nn.SoftMarginLoss()
    # criterion = WeightedMultilabel(weights=pos_weight)
    optimizer = optim.Adam(alexnet_model.parameters(), lr=lr, betas=(0.9, 0.99))
    # optimizer = optim.SGD(alexnet_model.parameters(), lr=lr, momentum=momentum, weight_decay=w_decay)
    # create dir for score
    score_dir = os.path.join(model_dir, 'scores')
    if not os.path.exists(score_dir):
        os.makedirs(score_dir)
    train_dict = {'loss': []}
    val_dict = {'loss': [], 'auc': []}
    logger1 = logger.Logger(LogPath)
    best_loss = 0
    Results = []  # 创建二位空矩阵
    # epochs = 39 + 1
    for i in range(7):
        Results.append([])
        for j in range(epochs + 1):
            Results[i].append([])

    for epoch in range(1, epochs):
        # if epoch == 2:
        #     break
        # print(val_dict['loss'][0])
        adjust_learning_rate(optimizer, epoch)
        train(alexnet_model, train_loader, epoch, train_dict, logger1, criterion, use_gpu)
        print("------------------------fold", fold, '------------------------------')
        print("------------------------epoch", epoch, '------------------------------')
        print("------------------------", 'auc_train', '------------------------------')
        auc_train, loss_train, gt_train, pr_train, pr_train0, pr_train1 = val_test(alexnet_model, train_loader)
        # print("------------------------", 'auc_val', '------------------------------')
        # auc_val, loss_val = val_test(alexnet_model,  val_loader)
        print("------------------------", 'auc_test', '------------------------------')
        auc_test, loss_test, gt_test, pr_test, pr_test0, pr_test1 = val_test(alexnet_model, test_loader)
        # train_path = train_loader.dataset.image_files
        # test_path = test_loader.dataset.image_files

        Results[0][0] = 'epoch'
        Results[1][0] = 'auc_train'
        Results[2][0] = 'loss_train'
        # Results[3][0] = 'auc_val'
        # Results[4][0] = 'loss_val'
        Results[5][0] = 'auc_test'
        Results[6][0] = 'loss_test'

        Results[0][epoch] = epoch
        Results[1][epoch] = auc_train
        Results[2][epoch] = loss_train
        # Results[3][epoch] = auc_val
        # Results[4][epoch] = loss_val
        Results[5][epoch] = auc_test
        Results[6][epoch] = loss_test

        # if auc_test > 0.8 and auc_test < 0.83 and auc_val > 0.8 and auc_val < 0.85:
        #     model_path = os.path.join(model_dir, str(auc_test)[:4] + '_best_mobilev2.pth')
        #     torch.save(alexnet_model, model_path)
        if auc_train > 0.75:
            model_path = os.path.join(model_dir, str(auc_train)[:4] + '_train_fold' + str(fold) + modelName)
            torch.save(alexnet_model, model_path)
            # 指定文件的路径
            path = probPath + str(auc_train)[:4] + '_train_fold' + str(fold) + '_epoch_' + str(epoch) + '.xls'
            workbook = openpyxl.Workbook()
            sheet0 = workbook.create_sheet(index=0)  # 创建sheet0
            # 循环写入数据
            sheet0.cell(1, 1).value = 'train_path'
            sheet0.cell(1, 2).value = 'gt_train'
            sheet0.cell(1, 3).value = 'pr_train'
            sheet0.cell(1, 4).value = 'pr_train0'
            sheet0.cell(1, 5).value = 'pr_train1'
            for i in range(len(train_path)):
                sheet0.cell(i + 2, 1).value = train_path[i]  # 写入数据
                sheet0.cell(i + 2, 2).value = gt_train[i]  # 写入数据
                sheet0.cell(i + 2, 3).value = pr_train[i]  # 写入数据
                sheet0.cell(i + 2, 4).value = pr_train0[i]  # 写入数据
                sheet0.cell(i + 2, 5).value = pr_train1[i]  # 写入数据
                # sheet0.cell(i + 1, j + 1).alignment = Alignment(horizontal='center', vertical='center')  # 居中对齐
            workbook.save(path)  # 保存文件

        if auc_test > 0.7:
            model_path = os.path.join(model_dir, str(auc_test)[:4] + '_test_fold' + str(fold) + modelName)
            torch.save(alexnet_model, model_path)

            # 指定文件的路径
            path = probPath + str(auc_test)[:4] + 'test_fold' + str(fold) + '_epoch_' + str(epoch) + '.xls'
            workbook = openpyxl.Workbook()
            sheet0 = workbook.create_sheet(index=0)  # 创建sheet0
            # 循环写入数据
            sheet0.cell(1, 1).value = 'test_path'
            sheet0.cell(1, 2).value = 'gt_test'
            sheet0.cell(1, 3).value = 'pr_test'
            sheet0.cell(1, 4).value = 'pr_test0'
            sheet0.cell(1, 5).value = 'pr_test1'
            for i in range(len(test_path)):
                sheet0.cell(i + 2, 1).value = test_path[i]  # 写入数据
                sheet0.cell(i + 2, 2).value = gt_test[i]  # 写入数据
                sheet0.cell(i + 2, 3).value = pr_test[i]  # 写入数据
                sheet0.cell(i + 2, 4).value = pr_test0[i]  # 写入数据
                sheet0.cell(i + 2, 5).value = pr_test1[i]  # 写入数据
                # sheet0.cell(i + 1, j + 1).alignment = Alignment(horizontal='center', vertical='center')  # 居中对齐
            workbook.save(path)

        # if auc_val > 0.73:
        #     model_path = os.path.join(model_dir, str(auc_val)[:4] + '_val_best_mobilev2.pth')
        #     torch.save(alexnet_model, model_path)
    #
    # 结果保存到excel
    # 将数据写入第 i 行，第 j 列
    f = xlwt.Workbook()
    sheet1 = f.add_sheet(u'sheet1', cell_overwrite_ok=False)  # 创建sheet
    for i in range(7):
        sheet1.write(i, 0, str(Results[i][0]))
        # for j in range(np.size(datas)):
        for j in range(epoch):
            sheet1.write(i, j + 1, Results[i][j + 1])  # 将data[j] 写入第i行j列excel2003最大列为256
    path = aucExcelPath + '/Results' + '_AUC_fold' + str(fold) + '.xls'
    f.save(path)

print('finished')
    # filename = 'Results.json'
    # with open(filename, 'w') as file_obj:
    #     json.dump(Results, file_obj)

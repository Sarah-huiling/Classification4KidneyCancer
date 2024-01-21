
#!/usr/bin/env python
# coding=utf-8
import random

import openpyxl
import torch
import glob
import numpy as np
# from huaxi_cf_dataset import MyDataset
# from huaxi_dataset import MyDataset
# from data_WSI_tumor1vs2 import MyDataset
from data_WSI_1v0 import MyDataset
from torch.utils.data import DataLoader
from torch.autograd import Variable
from sklearn import metrics as mt
# from skimage import io, color
from PIL import Image, ImageDraw, ImageFont
import os
import pylab as plt
import xlrd
import xlwt
import pandas as pd
# import csv
# import codecs
from sklearn.metrics import confusion_matrix


os.environ['CUDA_LAUNCH_BLOCKING'] = "0, 1"

def adjust_learning_rate(optimizer, epoch):
    """Sets the learning rate to the initial LR decayed by 10 every 30 epochs"""
    lr = 0.001 * (0.1 ** (epoch // 30))
    for param_group in optimizer.param_groups:
        param_group['lr'] = lr



def ODIR_Metrics(pred, target):

    th = 0.5
    gt = target.flatten()
    pr = pred.flatten()

    gt1 = gt[0::2]
    pr_neg = pr[0::2]  # pr_neg 预测为阴性的概率
    gt2 = gt[1::2]  # one-hot后的gt
    pr_pos = pr[1::2]   # pr_pos 预测为阳性的概率

    gt_prePob = []
    for i in range(len(gt2)):
        if gt2[i] == 1:
            gt_prePob.append(pr_pos[i])
        if gt2[i] == 0:
            gt_prePob.append(pr_neg[i])
    preLabel = np.zeros(len(gt2))
    preLabel[pr_pos > th] = 1

    print('=' * 20)
    print('gt2.shape', gt2.shape)
    print('pr2.shape', pr_pos.shape)
    # print('pr_pos.shape', len(pr_pos))
    # fpr, tpr, thresholds = mt.roc_curve(gt2, pr_pos, pos_label=1.0)  # 阳性的概率：以1类作为阳性，则输入预测为1的概率
    # roc_auc2 = mt.auc(fpr, tpr)
    kappa = mt.cohen_kappa_score(gt, pr > th)
    # print("1：auc值,", mt.roc_auc_score(gt1, pr_neg), 'acc:', mt.accuracy_score(gt1, pr_neg > th))
    # print("2：auc值,", mt.roc_auc_score(gt2, pr_pos), 'acc:', mt.accuracy_score(gt2, pr_pos > th))
    # # f1 = mt.f1_score(gt, pr > th, average='micro')
    # roc_auc = mt.roc_auc_score(gt2, pr_pos)
    fpr, tpr, thresholds = mt.roc_curve(gt2, pr_pos, pos_label=1.0)  # 阳性的概率：以1类作为阳性，则输入预测为1的概率
    roc_auc = mt.auc(fpr, tpr)
    acc=mt.accuracy_score(gt2, preLabel)
    print("auc:", roc_auc, 'acc:', acc)
    return roc_auc, gt2, gt_prePob, pr_neg, pr_pos,acc

def val_test(alexnet_model, val_loader, titlename):
    alexnet_model.eval()
    with torch.no_grad():
        p = []
        g = []
        PIDnames = []
        for iter, batch in enumerate(val_loader):
            torch.cuda.empty_cache()
            if use_gpu:
                inputs = Variable(batch[0].cuda())
                # cf = Variable(batch[1].cuda())
                labels = Variable(batch[1].cuda())
                # infos = Variable(batch[2].cuda())
                # file_name = batch[2]
            else:
                inputs, labels = Variable(batch['0']), Variable(batch['1'])
                # inputs, labels, infos = Variable(batch['0']), Variable(batch['1']), Variable(batch['2'])

            outputs = alexnet_model(inputs)
            # outputs = alexnet_model(inputs, infos)
            # loss = criterion(outputs, labels)
            outputs = torch.softmax(outputs, dim=1)
            outputs = outputs.data.cpu().numpy()
            labels = labels.cpu().numpy()
            # file_name = file_name.cpu().numpy()
            # PIDnames.append(file_name)
            for x, y in zip(outputs, labels):
                p.append(x)
                g.append(y)
        auc, gt2, pr2, pr_neg, pr_pos, acc = ODIR_Metrics(np.array(p), np.array(g))
        return auc, gt2, pr2, pr_neg, pr_pos, acc


if __name__ == "__main__":
    batch_size = 64  # 8  32
    use_gpu = torch.cuda.is_available()
    num_gpu = list(range(torch.cuda.device_count()))
    # random.shuffle(t_path)

    # test_path = glob.glob('/media/zhl/ResearchData/20221025Huaxi-KidneyCancer-WSI/HeatMap/VahadancePatches/FH-Z1824875-6Full/*.jpg')

    # dataPaths = pd.read_csv('/media/zhl/ResearchData/20221025Huaxi-KidneyCancer-WSI/HeatMap/testPath.csv')
    # test_path = []
    # for test_path1 in dataPaths.values:
    #     test_path.append(str(test_path1[0]))

    savePath = '/media/zhl/ResearchData/20221025Huaxi-KidneyCancer-WSI/DLResults/TCGA-publicData'
    test_path = []
    dataPaths = '/media/zhl/ResearchData/20221025Huaxi-KidneyCancer-WSI/patch/vahadance'
    for data in os.listdir(dataPaths):
        test_path += sorted(glob.glob(os.path.join(dataPaths, data) + '/*.jpg'))

    test = MyDataset(test_path, transform=False)
    test_loader = DataLoader(test, batch_size=batch_size, shuffle=False, num_workers=8)

    model_dir = "/media/zhl/ResearchData/20221025Huaxi-KidneyCancer-WSI/DLResults/Noverlap54PID/Vit_1vs0_patch32/model/best_model.pth"

    model = torch.load(model_dir)
    print(model)

    print('----------------------test value----------------------')
    auc_test, gt_test, pr_test, pr_test0, pr_test1, acc_test = val_test(model, test_loader, 'test')

    # 指定文件的路径
    # path = savePath + os.path.split(model_dir)[-1] + str(auc_test)[:5] + '_testProb.xls'
    path = os.path.join(savePath, 'acc_'+str(acc_test)[:5] + '_testProb1v0.xls')
    # path = savePath+str(auc_test)[:5] + '_testProb1v0.xls'
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0
    # 循环写入数据
    sheet0.cell(1, 1).value = 'test_path'
    sheet0.cell(1, 2).value = 'gt_test'
    sheet0.cell(1, 3).value = 'pr_test'
    sheet0.cell(1, 4).value = 'pr_test0'
    sheet0.cell(1, 5).value = 'pr_test1'
    for i in range(len(test_path)):
        sheet0.cell(i + 2, 1).value = test_path[i]
        sheet0.cell(i + 2, 2).value = gt_test[i]
        sheet0.cell(i + 2, 3).value = pr_test[i]
        sheet0.cell(i + 2, 4).value = pr_test0[i]
        sheet0.cell(i + 2, 5).value = pr_test1[i]
    workbook.save(path)
print('finish')


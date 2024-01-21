
# !/usr/bin/env python
# coding=utf-8
import random

import openpyxl
import torch
import glob
import numpy as np
# from huaxi_cf_dataset import MyDataset
# from huaxi_dataset import MyDataset
from data_WSI_tumor1vs2 import MyDataset
# from data_WSI import MyDataset
from torch.utils.data import DataLoader
from torch.autograd import Variable
from sklearn import metrics as mt
# from skimage import io, color
from PIL import Image, ImageDraw, ImageFont
import os
import pylab as plt
import xlrd
import xlwt
import pandas as pd
# import csv
# import codecs
from sklearn.metrics import confusion_matrix



os.environ['CUDA_LAUNCH_BLOCKING'] = "0, 1"


def adjust_learning_rate(optimizer, epoch):
    """Sets the learning rate to the initial LR decayed by 10 every 30 epochs"""
    lr = 0.001 * (0.1 ** (epoch // 30))
    for param_group in optimizer.param_groups:
        param_group['lr'] = lr


def draw_auc(fpr, tpr, name):
    roc_auc = mt.auc(fpr, tpr)
    plt.figure(figsize=(6, 6))
    tile = name + ' ROC'
    plt.title(tile)
    plt.plot(fpr, tpr, 'b', label=name + ' = %0.3f' % roc_auc)
    plt.legend(loc='lower right')
    plt.plot([0, 1], [0, 1], 'r--')
    plt.xlim([0, 1])
    plt.ylim([0, 1])
    plt.ylabel('True Positive Rate')
    plt.xlabel('False Positive Rate')
    plt.show()


def drawCM(matrix, savname):
    # Display different color for different elements
    lines, cols = matrix.shape
    sumline = matrix.sum(axis=1).reshape(lines, 1)
    ratiomat = matrix / sumline
    toplot0 = 1 - ratiomat
    toplot = toplot0.repeat(50).reshape(lines, -1).repeat(50, axis=0)
    io.imsave(savname, color.gray2rgb(toplot))
    # Draw values on every block
    image = Image.open(savname)
    draw = ImageDraw.Draw(image)
    font = ImageFont.truetype(os.path.join(os.getcwd(), "draw/ARIAL.TTF"), 15)
    for i in range(lines):
        for j in range(cols):
            dig = str(matrix[i, j])
            if i == j:
                filled = (255, 181, 197)
            else:
                filled = (46, 139, 87)
            draw.text((50 * j + 10, 50 * i + 10), dig, font=font, fill=filled)
    image.save(savname, 'jpeg')


def va(gt2, pr2, th):
    value_0 = {'tp': 0, 'tn': 0, 'fn': 0, 'fp': 0}
    for i in range(len(gt2)):
        if gt2[i] == 1 and pr2[i] >= th:
            value_0['tp'] = value_0['tp'] + 1  # 真正例
        if gt2[i] == 0 and pr2[i] >= th:
            value_0['fp'] = value_0['fp'] + 1  # 假正例
        if gt2[i] == 0 and pr2[i] < th:
            value_0['tn'] = value_0['tn'] + 1  # 真负例
        if gt2[i] == 1 and pr2[i] < th:
            value_0['fn'] = value_0['fn'] + 1  # 假负例
    return value_0


def ODIR_Metrics(pred, target):
    # corrected
    th = 0.5
    gt = target.flatten()
    pr = pred.flatten()

    gt1 = gt[0::2]
    pr_neg = pr[0::2]  # pr_neg 预测为阴性的概率
    gt2 = gt[1::2]
    pr_pos = pr[1::2]   # pr_pos 预测为阳性的概率

    gt_prePob = []
    for i in range(len(gt2)):
        if gt2[i] == 1:
            gt_prePob.append(pr_pos[i])
        if gt2[i] == 0:
            gt_prePob.append(pr_neg[i])
    preLabel = np.zeros(len(gt2))
    preLabel[pr_pos > th] = 1

    print('=' * 20)
    print('gt2.shape', gt2.shape)
    print('pr2.shape', pr_pos.shape)
    # print('pr_pos.shape', len(pr_pos))
    # fpr, tpr, thresholds = mt.roc_curve(gt2, pr_pos, pos_label=1.0)  # 阳性的概率：以1类作为阳性，则输入预测为1的概率
    # roc_auc2 = mt.auc(fpr, tpr)
    kappa = mt.cohen_kappa_score(gt, pr > th)
    # print("1：auc值,", mt.roc_auc_score(gt1, pr_neg), 'acc:', mt.accuracy_score(gt1, pr_neg > th))
    # print("2：auc值,", mt.roc_auc_score(gt2, pr_pos), 'acc:', mt.accuracy_score(gt2, pr_pos > th))
    # # f1 = mt.f1_score(gt, pr > th, average='micro')
    # roc_auc = mt.roc_auc_score(gt2, pr_pos)
    fpr, tpr, thresholds = mt.roc_curve(gt2, pr_pos, pos_label=1.0)  # 阳性的概率：以1类作为阳性，则输入预测为1的概率
    roc_auc = mt.auc(fpr, tpr)
    print("1：auc值,", roc_auc)
    return roc_auc, gt2, gt_prePob, pr_neg, pr_pos


def val_test(alexnet_model, val_loader, titlename):
    alexnet_model.eval()
    with torch.no_grad():
        p = []
        g = []
        PIDnames = []
        for iter, batch in enumerate(val_loader):
            torch.cuda.empty_cache()
            if use_gpu:
                inputs = Variable(batch[0].cuda())
                # cf = Variable(batch[1].cuda())
                labels = Variable(batch[1].cuda())
                # infos = Variable(batch[2].cuda())
                # file_name = batch[2]
            else:
                inputs, labels = Variable(batch['0']), Variable(batch['1'])
                # inputs, labels, infos = Variable(batch['0']), Variable(batch['1']), Variable(batch['2'])

            outputs = alexnet_model(inputs)
            # outputs = alexnet_model(inputs, infos)
            # loss = criterion(outputs, labels)
            outputs = torch.softmax(outputs, dim=1)
            outputs = outputs.data.cpu().numpy()
            labels = labels.cpu().numpy()
            # file_name = file_name.cpu().numpy()
            # PIDnames.append(file_name)
            for x, y in zip(outputs, labels):
                p.append(x)
                g.append(y)
        auc, gt2, pr2, pr_neg, pr_pos = ODIR_Metrics(np.array(p), np.array(g))
        return auc, gt2, pr2, pr_neg, pr_pos


if __name__ == "__main__":
    batch_size = 64  # 8  32
    use_gpu = torch.cuda.is_available()
    num_gpu = list(range(torch.cuda.device_count()))
    # random.shuffle(t_path)

    savePath = '/media/zhl/ResearchData/20221025Huaxi-KidneyCancer-WSI/20230511/patch-full/inferPrediction'
    dataPaths = '/media/zhl/ResearchData/20221025Huaxi-KidneyCancer-WSI/20230511/patch/vahadance'
    for data in os.listdir(dataPaths):
        test_path += sorted(glob.glob(os.path.join(dataPaths, data) + '/*.jpg'))

    test = MyDataset(test_path, transform=False)

    print('model load...')
    model_dir = "/media/zhl/ResearchData/20221025Huaxi-KidneyCancer-WSI/DLResults/Noverlap14PID/Vit_1vs2_vahadance/model/best_model.pth"
    model = torch.load(model_dir)
    print(model)
    print('----------------------test value----------------------')
    auc_test, gt_test, pr_test, pr_test0, pr_test1 = val_test(model, test_loader, 'test')

    # path = savePath + os.path.split(model_dir)[-1] + str(auc_test)[:5] + '_testProb.xls'
    path = os.path.join(savePath, PID + str(auc_test)[:5] + '_testProb1v2.xls')
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0
    # 循环写入数据
    sheet0.cell(1, 1).value = 'test_path'
    sheet0.cell(1, 2).value = 'gt_test'
    sheet0.cell(1, 3).value = 'pr_test'
    sheet0.cell(1, 4).value = 'pr_test0'
    sheet0.cell(1, 5).value = 'pr_test1'
    for i in range(len(test_path)):
        sheet0.cell(i + 2, 1).value = test_path[i]  # 写入数据
        sheet0.cell(i + 2, 2).value = gt_test[i]  # 写入数据
        sheet0.cell(i + 2, 3).value = pr_test[i]  # 写入数据
        sheet0.cell(i + 2, 4).value = pr_test0[i]  # 写入数据
        sheet0.cell(i + 2, 5).value = pr_test1[i]  # 写入数据
        # sheet0.cell(i + 1, j + 1).alignment = Alignment(horizontal='center', vertical='center')  # 居中对齐
    workbook.save(path)
print('finish')
